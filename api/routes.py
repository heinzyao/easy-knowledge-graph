"""
FastAPI routes: graph data, node preview, tag list, WebSocket updates.
Module-level globals (graph_builder, base_dir, ws_connections) are
injected by main.py at startup.
"""
from __future__ import annotations

import asyncio
import json
import urllib.request
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from fastapi import APIRouter, HTTPException, Query, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse

if TYPE_CHECKING:
    from core.graph_builder import GraphBuilder

router = APIRouter()

# Injected by main.py
graph_builder: GraphBuilder = None  # type: ignore[assignment]
base_dir: Path = None               # type: ignore[assignment]
ws_connections: list[WebSocket] = []


# ------------------------------------------------------------------
# REST endpoints
# ------------------------------------------------------------------

@router.get("/api/graph")
async def get_graph(include_implicit: bool = Query(default=False)) -> dict[str, Any]:
    return graph_builder.to_d3_format(include_implicit=include_implicit)


@router.get("/api/tags")
async def get_tags() -> list[str]:
    return graph_builder.get_all_tags()


@router.get("/api/node")
async def get_node(
    id: str = Query(..., description="Node ID (relative path)"),
    sheet: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=5000),
) -> dict[str, Any]:
    node = graph_builder.get_node_data(id)
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")

    file_path = base_dir / id
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        content_sheets = [s for s in wb.sheetnames if s != "Meta"]

        # Choose which sheet to preview
        target = sheet if (sheet and sheet in wb.sheetnames) else (content_sheets[0] if content_sheets else None)

        preview: dict[str, list] = {}
        total_rows = 0

        if target:
            ws = wb[target]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= limit:
                    break
                rows.append([("" if c is None else str(c)) for c in row])
            preview[target] = rows
            total_rows = ws.max_row or 0

        wb.close()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {
        "meta": {
            "id": node["id"],
            "title": node["title"],
            "tags": node["tags"],
            "links": node["links"],
            "description": node["description"],
        },
        "sheet_names": node["sheet_names"],
        "active_sheet": target,
        "preview": preview,
        "total_rows": total_rows,
        "limit": limit,
    }


@router.post("/api/refresh")
async def refresh_graph() -> dict[str, Any]:
    """Force a full rescan of the data directory."""
    graph_builder.build(include_implicit=True)
    await _broadcast({"type": "refresh"})
    return {"status": "ok", "node_count": graph_builder.node_count}


# ------------------------------------------------------------------
# WebSocket
# ------------------------------------------------------------------

@router.websocket("/ws/updates")
async def ws_updates(websocket: WebSocket) -> None:
    await websocket.accept()
    ws_connections.append(websocket)
    try:
        while True:
            # Keep connection alive; actual messages are pushed from the server
            await websocket.receive_text()
    except WebSocketDisconnect:
        pass
    finally:
        if websocket in ws_connections:
            ws_connections.remove(websocket)


async def _broadcast(message: dict[str, Any]) -> None:
    dead: list[WebSocket] = []
    for ws in list(ws_connections):
        try:
            await ws.send_json(message)
        except Exception:
            dead.append(ws)
    for ws in dead:
        if ws in ws_connections:
            ws_connections.remove(ws)


# Expose broadcast so main.py can call it from the watcher callback
broadcast = _broadcast


# ------------------------------------------------------------------
# Standalone HTML export
# ------------------------------------------------------------------

@router.get("/api/export")
async def export_standalone() -> HTMLResponse:
    """Generate a fully self-contained HTML file with all graph data embedded."""
    # 1. Graph data (include implicit so toggle works in standalone)
    graph_data = graph_builder.to_d3_format(include_implicit=True)

    # 2. Per-node preview data (meta + sheets)
    node_data: dict[str, Any] = {}
    for node in graph_data["nodes"]:
        node_id = node["id"]
        file_path = base_dir / node_id
        if not file_path.exists():
            continue
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            content_sheets = [s for s in wb.sheetnames if s != "Meta"]
            sheets: dict[str, Any] = {}
            for sheet_name in content_sheets:
                ws = wb[sheet_name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 200:
                        break
                    rows.append([("" if c is None else str(c)) for c in row])
                sheets[sheet_name] = {"rows": rows, "total_rows": ws.max_row or 0}
            wb.close()
            node_meta = graph_builder.get_node_data(node_id) or {}
            node_data[node_id] = {
                "meta": {
                    "id": node_id,
                    "title": node_meta.get("title", node_id),
                    "tags": node_meta.get("tags", []),
                    "links": node_meta.get("links", []),
                    "description": node_meta.get("description", ""),
                },
                "sheet_names": content_sheets,
                "sheets": sheets,
            }
        except Exception:
            pass

    # 3. Read frontend assets
    frontend_dir = Path(__file__).parent.parent / "frontend"
    style_css = (frontend_dir / "style.css").read_text(encoding="utf-8")
    graph_js = (frontend_dir / "graph.js").read_text(encoding="utf-8")
    panel_js = (frontend_dir / "panel.js").read_text(encoding="utf-8")

    # 4. Try to embed D3.js; fall back to CDN reference
    def _fetch_d3() -> str | None:
        try:
            with urllib.request.urlopen("https://d3js.org/d3.v7.min.js", timeout=15) as r:
                return r.read().decode("utf-8")
        except Exception:
            return None

    loop = asyncio.get_event_loop()
    d3_content = await loop.run_in_executor(None, _fetch_d3)
    if d3_content:
        d3_tag = f"<script>{d3_content}</script>"
    else:
        d3_tag = '<script src="https://d3js.org/d3.v7.min.js"></script>'

    # 5. Patch JS for standalone (no server required)
    standalone_graph_js = _patch_graph_js(graph_js)
    standalone_panel_js = _patch_panel_js(panel_js)

    # 6. Assemble HTML
    html = _build_export_html(graph_data, node_data, style_css,
                               standalone_graph_js, standalone_panel_js, d3_tag)
    return HTMLResponse(
        content=html,
        headers={"Content-Disposition": 'attachment; filename="knowledge-graph.html"'},
    )


def _safe_json(obj: Any) -> str:
    """JSON-encode and escape </script> to prevent HTML injection."""
    return json.dumps(obj, ensure_ascii=False).replace("</", "<\\/")


def _patch_graph_js(js: str) -> str:
    """Replace server-dependent functions for standalone mode."""
    # Replace loadGraph to use embedded data + support implicit toggle
    old_load = (
        "async function loadGraph() {\n"
        "    setLoading(true);\n"
        "    try {\n"
        "        const implicit = document.getElementById('implicit-toggle').checked;\n"
        "        const res = await fetch(`/api/graph?include_implicit=${implicit}`);\n"
        "        if (!res.ok) throw new Error(`HTTP ${res.status}`);\n"
        "        graphData = await res.json();\n"
        "        renderGraph();\n"
        "        updateTagFilter();\n"
        "        updateLegend();\n"
        "        updateNodeCount();\n"
        "    } catch (err) {\n"
        "        console.error('Failed to load graph:', err);\n"
        "    } finally {\n"
        "        setLoading(false);\n"
        "    }\n"
        "}"
    )
    new_load = (
        "async function loadGraph() {\n"
        "    setLoading(true);\n"
        "    try {\n"
        "        const implicit = document.getElementById('implicit-toggle').checked;\n"
        "        const all = window.__KG_GRAPH_DATA__;\n"
        "        graphData = implicit ? all : {\n"
        "            nodes: all.nodes,\n"
        "            links: all.links.filter(l => l.type !== 'implicit'),\n"
        "        };\n"
        "        renderGraph();\n"
        "        updateTagFilter();\n"
        "        updateLegend();\n"
        "        updateNodeCount();\n"
        "    } catch (err) {\n"
        "        console.error('Failed to load graph:', err);\n"
        "    } finally {\n"
        "        setLoading(false);\n"
        "    }\n"
        "}"
    )
    js = js.replace(old_load, new_load)

    # Disable WebSocket
    old_ws_start = "function initWebSocket() {\n    const proto = location.protocol"
    new_ws = "function initWebSocket() { /* standalone mode — no WebSocket */ }"
    # Find and replace the entire initWebSocket function
    ws_idx = js.find("function initWebSocket() {")
    if ws_idx != -1:
        brace_depth = 0
        i = js.index("{", ws_idx)
        for j in range(i, len(js)):
            if js[j] == "{":
                brace_depth += 1
            elif js[j] == "}":
                brace_depth -= 1
                if brace_depth == 0:
                    js = js[:ws_idx] + new_ws + js[j + 1:]
                    break

    # Disable refresh button (replace handler with disabled state)
    old_refresh = (
        "    document.getElementById('refresh-btn').addEventListener('click', async () => {\n"
        "        const btn = document.getElementById('refresh-btn');\n"
        "        btn.textContent = '⟳ 掃描中…';\n"
        "        btn.disabled = true;\n"
        "        try {\n"
        "            await fetch('/api/refresh', { method: 'POST' });\n"
        "            await loadGraph();\n"
        "        } finally {\n"
        "            btn.textContent = '⟳ 重新整理';\n"
        "            btn.disabled = false;\n"
        "        }\n"
        "    });"
    )
    new_refresh = (
        "    const _refreshBtn = document.getElementById('refresh-btn');\n"
        "    if (_refreshBtn) { _refreshBtn.disabled = true; _refreshBtn.title = '匯出版本不支援重新整理'; }"
    )
    js = js.replace(old_refresh, new_refresh)

    return js


def _patch_panel_js(js: str) -> str:
    """Replace API-fetch functions with embedded-data lookups."""
    # Replace window.openPanel
    old_open = (
        "window.openPanel = async function(nodeId) {\n"
        "    currentNodeId = nodeId;\n"
        "\n"
        "    const panel = document.getElementById('panel');\n"
        "    panel.classList.remove('panel-hidden');\n"
        "    document.getElementById('panel-title').textContent = '載入中…';\n"
        "    document.getElementById('panel-meta').innerHTML = '';\n"
        "    document.getElementById('panel-tabs').innerHTML = '';\n"
        "    document.getElementById('panel-table').innerHTML = '';\n"
        "    document.getElementById('panel-footer').textContent = '';\n"
        "\n"
        "    try {\n"
        "        const res = await fetch(`/api/node?id=${encodeURIComponent(nodeId)}`);\n"
        "        if (!res.ok) throw new Error(`HTTP ${res.status}`);\n"
        "        const data = await res.json();\n"
        "        renderPanel(data);\n"
        "    } catch (err) {\n"
        "        document.getElementById('panel-title').textContent = '載入失敗';\n"
        "        console.error('Panel load error:', err);\n"
        "    }\n"
        "};"
    )
    new_open = (
        "window.openPanel = async function(nodeId) {\n"
        "    currentNodeId = nodeId;\n"
        "\n"
        "    const panel = document.getElementById('panel');\n"
        "    panel.classList.remove('panel-hidden');\n"
        "    document.getElementById('panel-title').textContent = '載入中…';\n"
        "    document.getElementById('panel-meta').innerHTML = '';\n"
        "    document.getElementById('panel-tabs').innerHTML = '';\n"
        "    document.getElementById('panel-table').innerHTML = '';\n"
        "    document.getElementById('panel-footer').textContent = '';\n"
        "\n"
        "    const stored = (window.__KG_NODE_DATA__ || {})[nodeId];\n"
        "    if (!stored) {\n"
        "        document.getElementById('panel-title').textContent = '節點資料不存在';\n"
        "        return;\n"
        "    }\n"
        "    const firstSheet = stored.sheet_names[0] || null;\n"
        "    renderPanel({\n"
        "        meta: stored.meta,\n"
        "        sheet_names: stored.sheet_names,\n"
        "        active_sheet: firstSheet,\n"
        "        preview: firstSheet && stored.sheets[firstSheet]\n"
        "            ? { [firstSheet]: stored.sheets[firstSheet].rows } : {},\n"
        "        total_rows: firstSheet && stored.sheets[firstSheet]\n"
        "            ? stored.sheets[firstSheet].total_rows : 0,\n"
        "        limit: 200,\n"
        "    });\n"
        "};"
    )
    js = js.replace(old_open, new_open)

    # Replace loadSheet
    old_load_sheet = (
        "async function loadSheet(sheetName) {\n"
        "    if (!currentNodeId) return;\n"
        "\n"
        "    // Update active tab UI immediately\n"
        "    document.querySelectorAll('.tab-btn').forEach(btn => {\n"
        "        btn.classList.toggle('active', btn.dataset.sheet === sheetName);\n"
        "    });\n"
        "\n"
        "    document.getElementById('panel-table').innerHTML = '<tr><td style=\"color:var(--text-muted)\">載入中…</td></tr>';\n"
        "\n"
        "    try {\n"
        "        const res = await fetch(\n"
        "            `/api/node?id=${encodeURIComponent(currentNodeId)}&sheet=${encodeURIComponent(sheetName)}&limit=100`\n"
        "        );\n"
        "        if (!res.ok) throw new Error(`HTTP ${res.status}`);\n"
        "        const data = await res.json();\n"
        "\n"
        "        if (data.preview[sheetName]) {\n"
        "            renderTable(data.preview[sheetName]);\n"
        "        } else {\n"
        "            document.getElementById('panel-table').innerHTML = '';\n"
        "        }\n"
        "        renderFooter(data.total_rows, data.limit, sheetName);\n"
        "    } catch (err) {\n"
        "        document.getElementById('panel-table').innerHTML =\n"
        "            `<tr><td style=\"color:var(--danger)\">載入失敗: ${err.message}</td></tr>`;\n"
        "    }\n"
        "}"
    )
    new_load_sheet = (
        "async function loadSheet(sheetName) {\n"
        "    if (!currentNodeId) return;\n"
        "\n"
        "    document.querySelectorAll('.tab-btn').forEach(btn => {\n"
        "        btn.classList.toggle('active', btn.dataset.sheet === sheetName);\n"
        "    });\n"
        "\n"
        "    const stored = (window.__KG_NODE_DATA__ || {})[currentNodeId];\n"
        "    if (!stored || !stored.sheets[sheetName]) {\n"
        "        document.getElementById('panel-table').innerHTML = '';\n"
        "        return;\n"
        "    }\n"
        "    renderTable(stored.sheets[sheetName].rows);\n"
        "    renderFooter(stored.sheets[sheetName].total_rows, 200, sheetName);\n"
        "}"
    )
    js = js.replace(old_load_sheet, new_load_sheet)

    return js


def _build_export_html(
    graph_data: Any,
    node_data: Any,
    style_css: str,
    graph_js: str,
    panel_js: str,
    d3_tag: str,
) -> str:
    graph_json = _safe_json(graph_data)
    node_json = _safe_json(node_data)

    return f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel 知識圖譜（獨立版）</title>
    {d3_tag}
    <style>{style_css}</style>
</head>
<body>

<!-- Toolbar -->
<div id="toolbar">
    <div id="toolbar-left">
        <span id="app-title">📊 知識圖譜</span>
        <input type="text" id="search-input" placeholder="搜尋節點名稱…" autocomplete="off">
        <select id="tag-filter">
            <option value="">全部標籤</option>
        </select>
    </div>
    <div id="toolbar-right">
        <label class="toggle-label" title="顯示由儲存格值自動偵測的隱式連結">
            <input type="checkbox" id="implicit-toggle"> 隱式連結
        </label>
        <button id="refresh-btn" title="匯出版本不支援重新整理" disabled>⟳ 重新整理</button>
        <span id="node-count"></span>
    </div>
</div>

<!-- Main -->
<div id="main">

    <!-- Graph -->
    <div id="graph-container">
        <svg id="graph-svg">
            <defs>
                <marker id="arrow-explicit" markerWidth="8" markerHeight="6"
                        refX="12" refY="3" orient="auto">
                    <polygon points="0 0, 8 3, 0 6" fill="#555"/>
                </marker>
                <marker id="arrow-implicit" markerWidth="8" markerHeight="6"
                        refX="12" refY="3" orient="auto">
                    <polygon points="0 0, 8 3, 0 6" fill="#5a7090"/>
                </marker>
            </defs>
        </svg>

        <!-- Legend -->
        <div id="legend">
            <div id="legend-title">標籤色彩</div>
            <div id="legend-items"></div>
        </div>

        <!-- Loading -->
        <div id="loading">載入中…</div>

        <!-- Empty state -->
        <div id="empty-state" class="hidden">
            <div class="empty-icon">📂</div>
            <div>尚未找到任何 Excel 檔案</div>
            <div style="font-size:11px">請將 .xlsx 檔案放入資料夾後重新整理</div>
        </div>
    </div>

    <!-- Preview panel -->
    <div id="panel" class="panel-hidden">
        <div id="panel-header">
            <div id="panel-title">—</div>
            <button id="panel-close" title="關閉">✕</button>
        </div>
        <div id="panel-meta"></div>
        <div id="panel-tabs"></div>
        <div id="panel-table-container">
            <table id="panel-table"></table>
        </div>
        <div id="panel-footer"></div>
    </div>

</div>

<!-- Tooltip -->
<div id="tooltip"></div>

<script>
window.__KG_GRAPH_DATA__ = {graph_json};
window.__KG_NODE_DATA__ = {node_json};
</script>
<script>{graph_js}</script>
<script>{panel_js}</script>
</body>
</html>"""
