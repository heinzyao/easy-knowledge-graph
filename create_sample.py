"""
建立 Apple iPhone 供應鏈範例 Excel 檔案。
每個 .xlsx 代表一家真實/代表性供應商，連結關係構成 iPhone 供應鏈知識圖譜。
執行方式：uv run python create_sample.py
"""
import openpyxl
from pathlib import Path


def meta_sheet(wb, title: str, tags: str, links: str, description: str):
    ws = wb.active
    ws.title = "Meta"
    ws.append(["title",       title])
    ws.append(["tags",        tags])
    ws.append(["links",       links])
    ws.append(["description", description])
    return ws


def save(wb, folder: Path, name: str):
    wb.save(folder / f"{name}.xlsx")
    print(f"  建立：{name}.xlsx")


def main():
    folder = Path(__file__).parent / "sample_data"
    folder.mkdir(exist_ok=True)

    # 清除舊檔
    for old in folder.glob("*.xlsx"):
        old.unlink()

    # ── iPhone 供應鏈結構 ──────────────────────────────────────────
    #
    #  台積電（晶片代工）─────────────────────────────┐
    #  三星（DRAM/NAND）──────────────────────────────┤
    #  索尼（CMOS 感測器）──→ 富士康（組裝）          │
    #  康寧（玻璃）          ↑                        │
    #  村田製作所（被動元件）┘                        ↓
    #                                           Apple（品牌端）──→ UPS/FedEx（物流）
    #                                                             ↓（隱式）
    #                                                        Best Buy（零售，無 Meta）
    #
    # ── 1. 台積電（Tier 1 晶片代工）─────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="台積電",
        tags="半導體, 晶片代工, Tier-1",
        links="",
        description="全球最大晶圓代工廠，為 Apple A 系列與 M 系列晶片提供 3nm / 5nm 製程代工",
    )

    ws = wb.create_sheet("製程節點")
    ws.append(["製程", "電晶體密度 (M/mm²)", "首批量產年", "主要客戶", "iPhone 用途"])
    ws.append(["N3E (3nm)",  "171", "2023", "Apple",        "A17 Pro (iPhone 15 Pro)"])
    ws.append(["N4P (4nm)",  "133", "2023", "Apple, Qualcomm", "A16 (iPhone 15)"])
    ws.append(["N5  (5nm)",  "173", "2020", "Apple",        "A14 Bionic (iPhone 12)"])
    ws.append(["N7  (7nm)",   "91", "2018", "Apple, AMD",   "A13 Bionic (iPhone 11)"])

    ws2 = wb.create_sheet("年度產能")
    ws2.append(["年份", "12吋晶圓片/月 (K)", "Apple 佔比 (%)", "資本支出 (億 USD)"])
    ws2.append(["2022", "250", "26", "363"])
    ws2.append(["2023", "260", "25", "304"])
    ws2.append(["2024", "280", "27", "300"])
    ws2.append(["2025E","300", "28", "380"])

    save(wb, folder, "台積電")

    # ── 2. 三星（Tier 1 記憶體）──────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="三星電子",
        tags="記憶體, DRAM, NAND, Tier-1",
        links="",
        description="全球最大 DRAM 與 NAND Flash 製造商，供應 iPhone 的 LPDDR5 記憶體與 UFS 儲存晶片",
    )

    ws = wb.create_sheet("記憶體產品")
    ws.append(["型號", "類型", "製程", "容量", "iPhone 機型", "單價 (USD)"])
    ws.append(["LPDDR5X", "DRAM",      "12nm", "8GB",  "iPhone 15 Pro", "18.50"])
    ws.append(["LPDDR5",  "DRAM",      "14nm", "6GB",  "iPhone 15",     "13.20"])
    ws.append(["V-NAND",  "NAND Flash","6層",  "256GB","iPhone 15 Pro", "22.00"])
    ws.append(["V-NAND",  "NAND Flash","6層",  "512GB","iPhone 15 Pro", "38.00"])
    ws.append(["V-NAND",  "NAND Flash","6層",  "1TB",  "iPhone 15 Pro Max","68.00"])

    ws2 = wb.create_sheet("出貨紀錄")
    ws2.append(["季度", "DRAM 出貨 (億 GB)", "NAND 出貨 (EB)", "Apple 份額 (%)", "ASP 變化 (%)"])
    ws2.append(["2024 Q1", "85",  "18.2", "14", "+3.2"])
    ws2.append(["2024 Q2", "92",  "19.8", "15", "+1.8"])
    ws2.append(["2024 Q3", "101", "22.1", "16", "+4.5"])
    ws2.append(["2024 Q4", "112", "24.3", "18", "+5.1"])

    save(wb, folder, "三星電子")

    # ── 3. 索尼（Tier 1 相機感測器）──────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="索尼半導體",
        tags="感測器, CMOS, 相機, Tier-1",
        links="",
        description="全球市占率第一的 CMOS 影像感測器製造商，供應 iPhone 主相機與前鏡頭感測器",
    )

    ws = wb.create_sheet("感測器規格")
    ws.append(["型號", "像素", "感光面積 (mm²)", "光圈", "iPhone 用途", "年供量 (M pcs)"])
    ws.append(["IMX903", "48MP", "1/1.28", "f/1.78", "iPhone 15 Pro 主鏡",  "80"])
    ws.append(["IMX772", "12MP", "1/2.55", "f/2.2",  "iPhone 15 前鏡頭",    "120"])
    ws.append(["IMX858", "48MP", "1/2.0",  "f/2.2",  "iPhone 15 超廣角",    "80"])
    ws.append(["IMX803", "12MP", "1/3.5",  "f/2.8",  "iPhone 15 Pro 望遠",  "60"])

    ws2 = wb.create_sheet("技術藍圖")
    ws2.append(["世代", "技術重點", "量產年", "應用機型"])
    ws2.append(["Exmor RS 5G", "堆疊式 BSI + AI ISP",  "2023", "iPhone 15 系列"])
    ws2.append(["Exmor RS 6G", "2層電晶體像素結構",     "2025", "iPhone 17 系列"])
    ws2.append(["Exmor RS 7G", "有機光電轉換層",        "2027", "iPhone 19 系列（預計）"])

    save(wb, folder, "索尼半導體")

    # ── 4. 康寧（Tier 1 玻璃材料）────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="康寧",
        tags="玻璃, 材料, Tier-1",
        links="",
        description="Ceramic Shield 與 Gorilla Glass 製造商，供應 iPhone 螢幕保護玻璃與背板玻璃",
    )

    ws = wb.create_sheet("產品規格")
    ws.append(["產品", "硬度 (莫氏)", "厚度 (mm)", "跌落強度提升", "iPhone 用途"])
    ws.append(["Ceramic Shield Gen 2",  "7.5", "0.70", "4x vs 一般玻璃",  "iPhone 15 正面"])
    ws.append(["Gorilla Glass Victus 2","7.0", "0.55", "2x vs 前代",      "iPhone 15 背板"])
    ws.append(["Ceramic Shield Gen 1",  "7.5", "0.68", "4x vs 一般玻璃",  "iPhone 13/14 正面"])

    ws2 = wb.create_sheet("供貨合約")
    ws2.append(["客戶", "合約年", "品項", "年供量 (M 片)", "金額 (億 USD)"])
    ws2.append(["富士康", "2023–2025", "Ceramic Shield Gen 2",  "160", "4.8"])
    ws2.append(["富士康", "2023–2025", "Gorilla Glass Victus 2","120", "2.4"])

    save(wb, folder, "康寧")

    # ── 5. 村田製作所（Tier 1 被動元件）──────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="村田製作所",
        tags="被動元件, MLCC, 射頻, Tier-1",
        links="",
        description="全球最大 MLCC（積層陶瓷電容）與射頻模組製造商，每台 iPhone 使用逾 1,000 顆村田元件",
    )

    ws = wb.create_sheet("主要元件")
    ws.append(["元件", "規格", "每台用量", "主要功能", "單價 (USD)"])
    ws.append(["MLCC",           "0402 / 100nF", "850", "濾波、去耦合",   "0.002"])
    ws.append(["Wi-Fi 射頻模組", "Wi-Fi 6E",     "1",   "無線通訊",       "3.20"])
    ws.append(["Bluetooth 模組", "BT 5.3",       "1",   "短距通訊",       "1.80"])
    ws.append(["壓電蜂鳴器",     "SMD 1206",     "2",   "觸覺回饋",       "0.45"])
    ws.append(["SAW 濾波器",     "4G/5G Band",   "18",  "行動通訊濾波",   "0.15"])

    ws2 = wb.create_sheet("年度出貨")
    ws2.append(["年份", "MLCC 出貨 (億顆)", "RF 模組 (M pcs)", "Apple 佔比 (%)", "營收 (億 USD)"])
    ws2.append(["2022", "4200", "180", "28", "152"])
    ws2.append(["2023", "4500", "195", "29", "148"])
    ws2.append(["2024", "4800", "215", "30", "161"])

    save(wb, folder, "村田製作所")

    # ── 6. 富士康（Tier 0 代工組裝）──────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="富士康",
        tags="組裝, EMS, Tier-0",
        links="台積電, 三星電子, 索尼半導體, 康寧, 村田製作所",
        description="全球最大電子代工廠，負責 iPhone 約 70% 的最終組裝，鄭州廠高峰期日產逾 50 萬台",
    )

    ws = wb.create_sheet("組裝產線")
    ws.append(["廠區", "國家", "月產能 (萬台)", "員工數 (萬人)", "主力機種"])
    ws.append(["鄭州廠",   "中國", "1,500", "20", "iPhone 15 / Pro"])
    ws.append(["成都廠",   "中國",   "400",  "6", "iPad / MacBook"])
    ws.append(["清奈廠",   "印度",   "300",  "4", "iPhone 15"])
    ws.append(["浦那廠",   "印度",   "200",  "3", "iPhone 14（轉產）"])
    ws.append(["越南廠",   "越南",   "150",  "2", "AirPods / iPad"])

    ws2 = wb.create_sheet("品質指標")
    ws2.append(["指標", "目標值", "2024 Q3 實績", "備註"])
    ws2.append(["良率 (%)",         "99.5", "99.6", "整機組裝良率"])
    ws2.append(["DPPM",             "< 50", "38",   "百萬台中缺陷數"])
    ws2.append(["OTD (%)",          "98.0", "98.4", "準時交貨率"])
    ws2.append(["客訴回應 (hrs)",   "< 24", "18",   "Apple NPI 標準"])

    save(wb, folder, "富士康")

    # ── 7. Apple（品牌端）────────────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="Apple",
        tags="品牌, 設計, OEM客戶",
        links="富士康, UPS物流",
        description="iPhone 設計與品牌擁有者，主導晶片自研（A/M 系列）、供應鏈管理與全球銷售",
    )

    ws = wb.create_sheet("iPhone 產品線")
    ws.append(["機型", "發布年", "晶片", "起售價 (USD)", "全球銷量 (M)", "主要市場"])
    ws.append(["iPhone 15 Pro Max", "2023", "A17 Pro",    "1,199", "42", "美國、中國"])
    ws.append(["iPhone 15 Pro",     "2023", "A17 Pro",    "999",   "38", "美國、歐洲"])
    ws.append(["iPhone 15 Plus",    "2023", "A16 Bionic", "899",   "18", "美國、亞洲"])
    ws.append(["iPhone 15",         "2023", "A16 Bionic", "799",   "55", "全球"])
    ws.append(["iPhone SE 3",       "2022", "A15 Bionic", "429",   "15", "新興市場"])

    ws2 = wb.create_sheet("供應商管理")
    ws2.append(["供應商", "供應項目", "供應商評級", "年採購額 (億 USD)", "雙重來源"])
    ws2.append(["台積電",   "A系列晶片代工",     "戰略級", "210", "否（獨家）"])
    ws2.append(["三星電子", "DRAM / NAND",        "關鍵級", "78",  "是（SK海力士）"])
    ws2.append(["索尼半導體","CMOS 感測器",       "關鍵級", "45",  "是（LG Innotek）"])
    ws2.append(["康寧",     "Ceramic Shield",     "關鍵級", "24",  "否（獨家）"])
    ws2.append(["村田製作所","MLCC / RF 模組",    "關鍵級", "38",  "是（TDK）"])
    ws2.append(["富士康",   "iPhone 組裝（70%）", "戰略級", "580", "是（立訊精密）"])

    save(wb, folder, "Apple")

    # ── 8. UPS物流（第三方物流）──────────────────────────────────
    wb = openpyxl.Workbook()
    meta_sheet(
        wb,
        title="UPS物流",
        tags="物流, 航空快遞, 供應鏈",
        links="Apple",
        description="Apple 全球主要物流合作夥伴，負責 iPhone 新品發布期間從富士康到全球零售通路的空運配送",
    )

    ws = wb.create_sheet("航線資料")
    ws.append(["起點", "目的地", "運輸方式", "標準天數", "月均量", "備註"])
    ws.append(["上海浦東", "美國肯塔基（UPS Hub）", "包機空運", "1", "800 萬台/新品季", "iPhone 發布包機"])
    ws.append(["清奈",     "美國肯塔基（UPS Hub）", "定期空運", "2", "120 萬台/月",     "印度產線"])
    ws.append(["上海浦東", "法蘭克福",              "空運",     "1", "280 萬台/新品季", "歐洲市場"])
    ws.append(["上海浦東", "東京成田",              "空運",     "1", "150 萬台/新品季", "日本市場"])

    ws2 = wb.create_sheet("倉儲中心")
    ws2.append(["地點", "類型", "面積 (m²)", "日吞吐量 (萬件)", "溫控", "保稅"])
    ws2.append(["美國 路易維爾", "全球樞紐",   "250,000", "500", "否", "是"])
    ws2.append(["德國 科隆",     "歐洲樞紐",   "120,000", "220", "否", "是"])
    ws2.append(["香港",          "亞太集散",    "80,000", "180", "否", "是"])
    ws2.append(["新加坡",        "東南亞集散",  "55,000", "100", "是", "是"])

    save(wb, folder, "UPS物流")

    # ── 9. Best Buy（無 Meta → 隱式連結）─────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "銷售紀錄"
    ws.append(["期間", "品牌", "機型", "銷售量 (萬台)", "營收 (億 USD)", "市占 (%)"])
    ws.append(["2024 Q4", "Apple", "iPhone 15 Pro Max", "42", "50.4", "18"])
    ws.append(["2024 Q4", "Apple", "iPhone 15 Pro",     "38", "38.0", "16"])
    ws.append(["2024 Q4", "Apple", "iPhone 15",         "55", "43.9", "23"])
    ws.append(["2024 Q4", "Apple", "iPhone SE 3",       "12",  "5.2",  "5"])
    ws.append(["2024 Q4", "Apple", "iPhone 14",         "20", "14.0",  "8"])

    ws2 = wb.create_sheet("庫存狀態")
    ws2.append(["機型", "目前庫存 (台)", "在途庫存 (台)", "預估週轉天數", "補貨狀態"])
    ws2.append(["iPhone 15 Pro Max 256G", "12,500", "8,000",  "18", "正常"])
    ws2.append(["iPhone 15 Pro 128G",     "18,200", "10,000", "15", "正常"])
    ws2.append(["iPhone 15 128G",         "35,600", "20,000", "20", "充足"])
    ws2.append(["iPhone 15 Pro Max 1T",   "2,800",  "3,000",  "12", "偏緊"])

    # 隱式連結：儲存格直接包含其他節點名稱
    ws3 = wb.create_sheet("供應商關係")
    ws3.append(["關係類型", "公司名稱", "說明"])
    ws3.append(["主要供貨商", "Apple",   "iPhone 獨家零售合作夥伴"])
    ws3.append(["物流合作",  "UPS物流", "到店配送與退貨物流"])

    save(wb, folder, "Best Buy")

    print(f"\n完成！共建立 9 個範例 Excel 於 {folder}")
    print("\nApple iPhone 供應鏈層級：")
    print("  [Tier 1] 台積電（晶片代工）")
    print("  [Tier 1] 三星電子（DRAM/NAND）")
    print("  [Tier 1] 索尼半導體（CMOS 感測器）")
    print("  [Tier 1] 康寧（Ceramic Shield 玻璃）")
    print("  [Tier 1] 村田製作所（MLCC/RF 元件）")
    print("  [Tier 0] 富士康（最終組裝，← 以上所有 Tier-1）")
    print("  [品牌]   Apple（← 富士康、UPS物流）")
    print("  [物流]   UPS物流（← Apple）")
    print("  [零售]   Best Buy（隱式連結 Apple、UPS物流）")
    print("\n執行：uv run python main.py  即可啟動")


if __name__ == "__main__":
    main()
